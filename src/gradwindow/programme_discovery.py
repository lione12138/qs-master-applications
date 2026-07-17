from __future__ import annotations

import hashlib
import json
from collections.abc import Callable
from datetime import date, datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .candidate_review import attach_programme_candidate_evidence_hash
from .http_client import DEFAULT_USER_AGENT, fetch_page
from .io import read_json, write_json
from .monitor import extract_fetched_text
from .paths import (
    APPLICATIONS_PATH,
    PROGRAMME_CANDIDATES_PATH,
    PROGRAMME_CATALOG_STATE_PATH,
    PROGRAMS_PATH,
    WINDOW_CANDIDATES_PATH,
)
from .predictions import official_cycle_key
from .programme_adapters.base import ProgrammeAdapter
from .programme_windows import (
    has_official_exact_window,
    known_programme_window_candidates,
)


def fetch_catalog(url: str) -> str:
    page = fetch_page(
        url,
        user_agent=DEFAULT_USER_AGENT,
        timeout=30,
        accept=(
            "text/html,application/xhtml+xml,application/xml;q=0.9,"
            "text/xml;q=0.8,*/*;q=0.7"
        ),
    )
    if (
        "spreadsheetml.sheet" in page.content_type.lower()
        or page.final_url.lower().split("?", 1)[0].endswith(".xlsx")
    ):
        return _xlsx_payload(page.raw_bytes)
    return extract_fetched_text(page)


def _xlsx_payload(raw_bytes: bytes) -> str:
    workbook = load_workbook(BytesIO(raw_bytes), read_only=True, data_only=True)
    worksheets = []
    for sheet in workbook.worksheets:
        rows = [
            [_json_cell(value) for value in row]
            for row in sheet.iter_rows(values_only=True)
        ]
        worksheets.append({"name": sheet.title, "rows": rows})
    workbook.close()
    return json.dumps({"worksheets": worksheets}, ensure_ascii=False)


def _json_cell(value):
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def discover_programmes(
    adapter: ProgrammeAdapter,
    *,
    programs_path: Path = PROGRAMS_PATH,
    applications_path: Path = APPLICATIONS_PATH,
    candidates_path: Path = PROGRAMME_CANDIDATES_PATH,
    window_candidates_path: Path | None = None,
    state_path: Path = PROGRAMME_CATALOG_STATE_PATH,
    fetcher: Callable[[str], str] = fetch_catalog,
    dry_run: bool = False,
) -> dict:
    checked_at = datetime.now(timezone.utc).isoformat()
    catalog = adapter.parse_catalog_from_fetcher(fetcher)
    programs_payload = read_json(programs_path)
    known_programmes = {
        item["id"]: item
        for item in programs_payload.get("programs", [])
        if item.get("universityId") == adapter.university_id
    }
    known_ids = set(known_programmes)
    candidates_payload = read_json(
        candidates_path,
        {
            "meta": {
                "description": (
                    "Automatically discovered programme candidates awaiting "
                    "manual review. This file is not published."
                )
            },
            "items": [],
        },
    )
    existing = {item["id"]: item for item in candidates_payload.get("items", [])}
    if adapter.replace_pending_candidates:
        existing = {
            candidate_id: item
            for candidate_id, item in existing.items()
            if item.get("universityId") != adapter.university_id
            or item.get("status", "pending") != "pending"
        }
    if window_candidates_path is None:
        window_candidates_path = (
            WINDOW_CANDIDATES_PATH
            if candidates_path == PROGRAMME_CANDIDATES_PATH
            else candidates_path.with_name("window-candidates.json")
        )
    applications = read_json(applications_path, {"applications": []}).get(
        "applications", []
    )
    window_candidates_payload = read_json(
        window_candidates_path,
        {
            "meta": {
                "description": (
                    "Internal exact-window candidates awaiting manual review. "
                    "This file is never published to the static site."
                )
            },
            "items": [],
        },
    )
    existing_window_candidates = {
        item["id"]: item for item in window_candidates_payload.get("items", [])
    }
    original_window_candidate_items = window_candidates_payload.get("items", [])
    applications_by_cycle = {
        official_cycle_key(item): item
        for item in applications
        if item.get("universityId") == adapter.university_id
    }
    application_ids = {item["id"] for item in applications}
    created = 0
    created_guidance_candidates = 0
    created_window_candidates = 0
    changed_window_candidates = 0
    for programme in catalog.programmes:
        if programme.id in known_ids:
            for candidate in known_programme_window_candidates(
                adapter,
                programme,
                known_programmes[programme.id],
                catalog.application_opens_at,
                applications_by_cycle,
                application_ids,
                checked_at,
            ):
                previous = existing_window_candidates.get(candidate["id"])
                if previous is not None:
                    candidate["status"] = previous.get("status", "pending")
                    candidate["detectedAt"] = previous.get("detectedAt", checked_at)
                    candidate["record"]["verifiedAt"] = previous.get("record", {}).get(
                        "verifiedAt", candidate["record"]["verifiedAt"]
                    )
                    for key in ("reviewedAt", "reviewedBy", "reviewNotes"):
                        if key in previous:
                            candidate[key] = previous[key]
                else:
                    created_window_candidates += 1
                    if candidate["type"] == "adapter-window-change":
                        changed_window_candidates += 1
                existing_window_candidates[candidate["id"]] = candidate
            guidance = _known_programme_guidance_candidate(
                adapter,
                programme,
                catalog.application_opens_at,
                checked_at,
            )
            if guidance is not None:
                previous = existing.get(guidance["id"])
                if previous is None:
                    created_guidance_candidates += 1
                else:
                    _merge_candidate_review_state(guidance, previous, checked_at)
                existing[guidance["id"]] = guidance
            continue
        candidate_id = f"new-programme:{programme.id}"
        previous = existing.get(candidate_id)
        candidate = _candidate_record(
            adapter,
            programme,
            catalog.application_opens_at,
            checked_at,
        )
        if previous is None:
            created += 1
        else:
            _merge_candidate_review_state(candidate, previous, checked_at)
        existing[candidate_id] = candidate

    items = sorted(
        existing.values(),
        key=lambda item: (item.get("status") != "pending", item["id"]),
    )
    snapshot_items = {
        programme.id: {
            "name": programme.name,
            "degreeType": programme.degree_type,
            "faculty": programme.faculty,
            "department": programme.department,
            "parseStatus": programme.parse_status,
            "deadlineHash": _hash(
                json.dumps(
                    [
                        {
                            "round": window.round,
                            "opensAt": window.opens_at,
                            "closesAt": window.closes_at,
                            "applicantCategories": window.applicant_categories,
                            "sourceUrl": window.source_url,
                        }
                        for window in programme.windows
                    ],
                    sort_keys=True,
                )
            ),
        }
        for programme in catalog.programmes
    }
    state_payload = read_json(state_path, {"meta": {}, "universities": {}})
    state_payload.setdefault("universities", {})[adapter.university_id] = {
        "sourceUrl": adapter.catalog_url,
        "checkedAt": checked_at,
        "itemCount": len(catalog.programmes),
        "applicationOpensAt": catalog.application_opens_at,
        "catalogHash": _hash(json.dumps(snapshot_items, sort_keys=True)),
        "programmes": snapshot_items,
    }
    state_payload["meta"] = {
        "updatedAt": checked_at,
        "description": "Latest official programme-catalog discovery snapshots.",
    }

    if not dry_run:
        candidates_payload["items"] = items
        candidates_payload.setdefault("meta", {})["updatedAt"] = checked_at
        write_json(candidates_path, candidates_payload)
        window_candidate_items = sorted(
            existing_window_candidates.values(),
            key=lambda item: (item.get("status") != "pending", item["id"]),
        )
        if window_candidate_items != original_window_candidate_items:
            window_candidates_payload["items"] = window_candidate_items
            window_candidates_payload.setdefault("meta", {})["updatedAt"] = checked_at
            write_json(window_candidates_path, window_candidates_payload)
        write_json(state_path, state_payload)

    return {
        "status": "ok",
        "universityId": adapter.university_id,
        "sourceUrl": adapter.catalog_url,
        "checkedAt": checked_at,
        "catalogProgrammes": len(catalog.programmes),
        "knownProgrammes": len(known_ids),
        "newCandidates": created,
        "newGuidanceCandidates": created_guidance_candidates,
        "newWindowCandidates": created_window_candidates,
        "changedWindowCandidates": changed_window_candidates,
        "pendingWindowCandidates": sum(
            item.get("status", "pending") == "pending"
            and item.get("universityId") == adapter.university_id
            for item in existing_window_candidates.values()
        ),
        "pendingCandidates": sum(
            item.get("status", "pending") == "pending"
            and item.get("universityId") == adapter.university_id
            for item in items
        ),
        "pendingGuidanceCandidates": sum(
            item.get("status", "pending") == "pending"
            and item.get("universityId") == adapter.university_id
            and item.get("type") == "known-programme-window-guidance"
            for item in items
        ),
        "programmesWithoutDeadlines": sum(
            not programme.windows for programme in catalog.programmes
        ),
        "programmesNeedingReview": sum(
            programme.parse_status != "parsed" for programme in catalog.programmes
        ),
        "dryRun": dry_run,
    }


def _known_programme_guidance_candidate(
    adapter,
    programme,
    shared_opens_at: str | None,
    detected_at: str,
) -> dict | None:
    candidate = _candidate_record(
        adapter,
        programme,
        shared_opens_at,
        detected_at,
    )
    unresolved_windows = [
        window
        for window in candidate["windows"]
        if not has_official_exact_window(window)
    ]
    if candidate["windows"] and not unresolved_windows:
        return None
    candidate["id"] = f"known-programme-guidance:{programme.id}"
    candidate["type"] = "known-programme-window-guidance"
    candidate["windows"] = unresolved_windows
    return attach_programme_candidate_evidence_hash(candidate)


def _candidate_record(
    adapter,
    programme,
    shared_opens_at: str | None,
    detected_at: str,
) -> dict:
    shared_opening_basis = adapter.application_opens_at_basis

    def opening_for(window) -> tuple[str | None, str]:
        opens_at = window.opens_at or shared_opens_at
        if not opens_at or opens_at > window.closes_at:
            return None, "missing"
        if window.opens_at:
            return opens_at, "official"
        return opens_at, shared_opening_basis

    windows = []
    for window in programme.windows:
        opens_at, opens_at_basis = opening_for(window)
        windows.append(
            {
                "intake": window.intake or adapter.intake,
                "round": window.round,
                "applicantCategories": window.applicant_categories,
                "opensAt": opens_at,
                "opensAtBasis": opens_at_basis,
                "closesAt": window.closes_at,
                "sourceUrl": window.source_url or programme.source_url,
            }
        )
    has_unresolved_opening = any(window["opensAt"] is None for window in windows)
    has_inferred_opening = any(
        window.get("opensAtBasis", "").startswith("inferred") for window in windows
    )
    deadline_precedes_shared_opening = bool(
        shared_opens_at
        and any(window["closesAt"] < shared_opens_at for window in windows)
    )
    candidate = {
        "id": f"new-programme:{programme.id}",
        "type": "new-programme",
        "status": "pending",
        "universityId": adapter.university_id,
        "detectedAt": detected_at,
        "sourceUrl": programme.source_url,
        "programme": {
            "id": programme.id,
            "universityId": adapter.university_id,
            "name": programme.name,
            "degreeType": programme.degree_type,
            "faculty": " | ".join(
                value for value in (programme.faculty, programme.department) if value
            ),
            "applicationUrl": programme.application_url,
            "sourceUrl": programme.source_url,
        },
        "windows": windows,
        "parseStatus": programme.parse_status,
        "reviewReason": (
            "No application deadline was parsed."
            if not windows
            else (
                (
                    "An early deadline precedes the shared commencement date; "
                    "confirm the programme-specific opening date."
                )
                if deadline_precedes_shared_opening
                else (
                    "At least one opening date is not published as an exact date; "
                    "confirm it on the programme page."
                )
                if has_unresolved_opening
                else (
                    "Opening date uses a configured cycle default; review the "
                    "officially parsed deadline before promotion."
                )
                if has_inferred_opening
                else "Review the automatically discovered programme and application rounds."
            )
        ),
        "evidenceExcerpt": programme.deadline_text,
    }
    if programme.retrieval_method or programme.evidence_quality:
        candidate["discoveryEvidence"] = {
            "retrievalMethod": programme.retrieval_method,
            "evidenceQuality": programme.evidence_quality,
            "documentHash": programme.evidence_document_hash,
        }
    return attach_programme_candidate_evidence_hash(candidate)


def _merge_candidate_review_state(
    candidate: dict[str, Any],
    previous: dict[str, Any],
    checked_at: str,
) -> None:
    previous_hash = previous.get("evidenceHash")
    evidence_changed = bool(
        previous_hash and previous_hash != candidate.get("evidenceHash")
    )
    candidate["detectedAt"] = (
        checked_at if evidence_changed else previous.get("detectedAt", checked_at)
    )
    candidate["status"] = (
        "pending" if evidence_changed else previous.get("status", "pending")
    )
    if previous.get("reviewHistory"):
        candidate["reviewHistory"] = previous["reviewHistory"]
    if evidence_changed:
        candidate["evidenceChangedAt"] = checked_at
        return
    for key in ("reviewedAt", "reviewedBy", "reviewNotes", "evidenceChangedAt"):
        if key in previous:
            candidate[key] = previous[key]


def _hash(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()
